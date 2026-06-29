"""
Consolidador de Horários - HORARIO JAGUARI 2026
Cria a aba "Horario Jaguari" consolidando todas as outras abas com formatação preservada.
"""

import copy
import os
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.styles.numbers import FORMAT_TEXT
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string

# ── Caminhos ──────────────────────────────────────────────────────────────────
BASE_DIR = r"c:\Users\Usuario1\Downloads\sistema-escolar-v2-main (2)\sistema-escolar-v2-main\Horario Jaguari 2026"
INPUT_FILE  = os.path.join(BASE_DIR, "HORARIO JAGUARI 2026 OFICIAL 09 02 26.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "HORARIO_JAGUARI_2026_CONSOLIDADO.xlsx")

NEW_SHEET_NAME = "Horario Jaguari"

# Cor da faixa divisória entre abas
DIVIDER_FILL  = PatternFill("solid", fgColor="1F3864")   # azul escuro
DIVIDER_FONT  = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
DIVIDER_ALIGN = Alignment(horizontal="center", vertical="center")

# ── Helpers ───────────────────────────────────────────────────────────────────

def copy_font(src: Font | None) -> Font:
    if src is None:
        return Font()
    return Font(
        name=src.name, size=src.size, bold=src.bold, italic=src.italic,
        underline=src.underline, strike=src.strike, color=copy.copy(src.color),
        vertAlign=src.vertAlign, charset=src.charset, scheme=src.scheme,
        family=src.family
    )

def copy_fill(src) -> PatternFill:
    try:
        if src is None:
            return PatternFill()
        fill_type = getattr(src, 'fill_type', None)
        if fill_type is None or fill_type == 'none':
            return PatternFill()
        if isinstance(src, PatternFill):
            fg = copy.copy(src.fgColor) if src.fgColor else None
            bg = copy.copy(src.bgColor) if src.bgColor else None
            return PatternFill(fill_type=fill_type, fgColor=fg, bgColor=bg)
        # GradientFill – converte para sólido usando a primeira parada
        if isinstance(src, GradientFill):
            stops = getattr(src, 'stop', None)
            if stops:
                return PatternFill('solid', fgColor=copy.copy(stops[0].value))
        return PatternFill()
    except Exception:
        return PatternFill()

def copy_border(src: Border | None) -> Border:
    if src is None:
        return Border()
    def cp_side(s):
        if s is None:
            return Side()
        return Side(style=s.style, color=copy.copy(s.color) if s.color else None)
    return Border(
        left=cp_side(src.left), right=cp_side(src.right),
        top=cp_side(src.top),   bottom=cp_side(src.bottom),
        diagonal=cp_side(src.diagonal),
        diagonalUp=src.diagonalUp, diagonalDown=src.diagonalDown
    )

def copy_alignment(src: Alignment | None) -> Alignment:
    if src is None:
        return Alignment()
    return Alignment(
        horizontal=src.horizontal, vertical=src.vertical,
        text_rotation=src.text_rotation, wrap_text=src.wrap_text,
        shrink_to_fit=src.shrink_to_fit, indent=src.indent,
        readingOrder=getattr(src, 'readingOrder', None)
    )

def copy_cell_style(src_cell, dst_cell):
    """Copia todo o estilo de uma célula para outra."""
    dst_cell.font      = copy_font(src_cell.font)
    dst_cell.fill      = copy_fill(src_cell.fill)
    dst_cell.border    = copy_border(src_cell.border)
    dst_cell.alignment = copy_alignment(src_cell.alignment)
    if src_cell.number_format:
        dst_cell.number_format = src_cell.number_format

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"Abrindo: {INPUT_FILE}")
    wb = load_workbook(INPUT_FILE)

    # Remove a aba consolidada se já existir
    if NEW_SHEET_NAME in wb.sheetnames:
        del wb[NEW_SHEET_NAME]
        print(f"Aba '{NEW_SHEET_NAME}' anterior removida.")

    # Cria a nova aba no final
    ws_new = wb.create_sheet(NEW_SHEET_NAME)
    print(f"Aba '{NEW_SHEET_NAME}' criada.")

    # Lista de abas de origem (excluindo a nova)
    source_sheets = [s for s in wb.sheetnames if s != NEW_SHEET_NAME]
    print(f"Abas de origem: {source_sheets}")

    current_row = 1          # linha de escrita na aba consolidada
    col_widths  = {}         # col_letter -> max width visto até agora

    for sheet_name in source_sheets:
        ws_src = wb[sheet_name]

        # Determina os limites reais da aba de origem
        max_row = ws_src.max_row
        max_col = ws_src.max_column
        if max_row is None or max_row == 0:
            print(f"  Aba '{sheet_name}' vazia – ignorada.")
            continue

        print(f"  Copiando '{sheet_name}': {max_row} linhas × {max_col} cols")

        # ── 1. Faixa divisória ──────────────────────────────────────────────
        div_row = current_row
        for c in range(1, max_col + 1):
            cell = ws_new.cell(row=div_row, column=c)
            cell.fill      = DIVIDER_FILL
            cell.font      = DIVIDER_FONT
            cell.alignment = DIVIDER_ALIGN

        # Coloca o nome da aba na primeira célula da faixa divisória
        ws_new.cell(row=div_row, column=1).value = f"◄ {sheet_name} ►"
        # Mescla a linha divisória ao longo de todas as colunas
        if max_col > 1:
            ws_new.merge_cells(
                start_row=div_row, start_column=1,
                end_row=div_row,   end_column=max_col
            )
        ws_new.row_dimensions[div_row].height = 20
        current_row += 1

        # ── 2. Coleta merges da aba de origem ──────────────────────────────
        # merge_map: (row, col) -> (min_row, min_col, max_row, max_col)
        merge_map = {}
        for mc in ws_src.merged_cells.ranges:
            for r in range(mc.min_row, mc.max_row + 1):
                for c in range(mc.min_col, mc.max_col + 1):
                    merge_map[(r, c)] = (mc.min_row, mc.min_col, mc.max_row, mc.max_col)

        # ── 3. Copia células ───────────────────────────────────────────────
        offset = current_row - 1   # deslocamento de linhas

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                src_cell = ws_src.cell(row=r, column=c)
                dst_cell = ws_new.cell(row=r + offset, column=c)
                dst_cell.value = src_cell.value
                copy_cell_style(src_cell, dst_cell)

        # ── 4. Replica merges ──────────────────────────────────────────────
        applied_merges = set()
        for mc in ws_src.merged_cells.ranges:
            key = (mc.min_row, mc.min_col, mc.max_row, mc.max_col)
            if key not in applied_merges:
                ws_new.merge_cells(
                    start_row=mc.min_row + offset, start_column=mc.min_col,
                    end_row=mc.max_row + offset,   end_column=mc.max_col
                )
                applied_merges.add(key)

        # ── 5. Alturas de linhas ───────────────────────────────────────────
        for r in range(1, max_row + 1):
            src_dim = ws_src.row_dimensions.get(r)
            if src_dim and src_dim.height:
                ws_new.row_dimensions[r + offset].height = src_dim.height

        # ── 6. Larguras de colunas ─────────────────────────────────────────
        for c in range(1, max_col + 1):
            col_letter = get_column_letter(c)
            src_dim = ws_src.column_dimensions.get(col_letter)
            w = src_dim.width if src_dim and src_dim.width else 8.43
            if col_letter not in col_widths or w > col_widths[col_letter]:
                col_widths[col_letter] = w

        current_row += max_row  # avança para além do bloco copiado

    # Aplica as larguras de coluna acumuladas
    for col_letter, width in col_widths.items():
        ws_new.column_dimensions[col_letter].width = width

    # Congela a primeira linha (opcional)
    # ws_new.freeze_panes = "A2"

    print(f"\nSalvando em: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print("[OK] Arquivo salvo com sucesso!")
    print(f"   Linhas totais na aba consolidada: {current_row - 1}")

if __name__ == "__main__":
    main()
