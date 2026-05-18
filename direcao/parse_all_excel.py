import openpyxl
import json
import re

file_path = r'c:\Users\Usuario1\Downloads\sistema-escolar-v2-main (2)\sistema-escolar-v2-main\Horario Jaguari 2026\HORARIO JAGUARI 2026 OFICIAL 09 02 26.xlsx'

wb = openpyxl.load_workbook(file_path, data_only=True)

# Mapeamento de aba -> chave JS
ID_MAP = {
    '1º A': '1A', '1ºB': '1B', '1º C': '1C',
    '2ºA': '2A',  '2ºB': '2B', '2º C': '2C',
    '3ºA': '3A',  '3º B': '3B', '3ºC': '3C',
    '4º A': '4A', '4º B': '4B', '4º C': '4C',
    '5ºA': '5A',  '5º B': '5B', '5ºC': '5C', '5º D': '5D',
    'EDILENE': 'EDILENE', 'MARCOS': 'MARCOS', 'MARJORIE': 'MARJORIE',
    'MIRIAM': 'MIRIAM', 'ARTES 1º ANO': 'ARTES1ANO', 'INGLÊS': 'INGLS',
    'OF. LEITURA': 'OFLEITURA', 'OF. MAKER': 'OFMAKER', 'OF. SEBRAE': 'OFSEBRAE',
    ' PA - MARCIA': 'PAMARCIA', ' PA  - DIRCEU ': 'PADIRCEU',
    'PA  - LOURDES ': 'PALOURDES', ' PA  - PIPOCA ': 'PAPIPOCA'
}

# Horários esperados para cada linha de aula (em ordem)
HORA_LABELS = [
    "7h30–8h20",
    "8h20–9h10",
    "9h10–9h30",   # intervalo
    "9h30–10h20",
    "10h20–11h10",
    "11h10–12h",
    "12h–13h",     # almoço
    "13h–13h50",
    "13h50–14h40",
    "14h40–15h",   # saída
    "15h–16h",
    "16h–17h",
    "17h–18h",
]

TIPOS_ESPECIAIS = {
    "9h10–9h30":  ("intervalo", "INTERVALO"),
    "12h–13h":    ("almoco",    "ALMOÇO"),
}

# Colunas dos dias: B=2, C=3, D=4, E=5, F=6
DAY_COLS = [2, 3, 4, 5, 6]


def get_cell_value(ws, row, col):
    """Lê o valor de uma célula resolvendo mescladas."""
    val = ws.cell(row=row, column=col).value
    if val is not None:
        return val
    for merged in ws.merged_cells.ranges:
        if (merged.min_row <= row <= merged.max_row and
                merged.min_col <= col <= merged.max_col):
            return ws.cell(row=merged.min_row, column=merged.min_col).value
    return None


def clean_cell(val):
    """Normaliza o texto de uma célula."""
    if val is None:
        return "—"
    s = str(val).strip()
    # Compacta múltiplos espaços
    s = re.sub(r'\s{2,}', ' ', s)
    if not s or s.lower() == 'none':
        return "—"
    return s


def is_intervalo_row(val_col_a, val_col_b):
    """Detecta se a linha é INTERVALO ou ALMOÇO pelo valor na col B."""
    if val_col_b and 'INTERVALO' in str(val_col_b).upper():
        return True
    if val_col_b and ('ALMOÇO' in str(val_col_b).upper() or 'ALMOCO' in str(val_col_b).upper()):
        return True
    return False


def detect_start_row(ws):
    """
    Detecta automaticamente a linha onde começam os dados de aula.
    Procura a linha que contém 'SEGUNDA' na coluna B (cabeçalho dos dias).
    A linha de dados começa na linha seguinte.
    """
    for r in range(1, 20):
        val = ws.cell(row=r, column=2).value
        if val and 'SEGUNDA' in str(val).upper():
            return r + 1  # linha de dados começa após o cabeçalho
    return 7  # fallback padrão


def detect_header_row(ws):
    """
    Detecta a linha do cabeçalho turma+prof.
    Procura nas primeiras linhas por um texto que contenha 'ANO' ou 'ED.' ou 'PROF' ou 'P.A.' etc.
    """
    for r in range(1, 8):
        val = ws.cell(row=r, column=1).value
        if val is None:
            continue
        s = str(val).strip()
        if s and not 'CIEP' in s.upper() and not 'HORÁRIO' in s.upper() and not 'HORARIO' in s.upper() and len(s) > 3:
            return r
    return 4  # fallback


def parse_turma_prof(raw):
    """Extrai turma e prof do texto do cabeçalho."""
    raw = str(raw).strip()
    # Divide pelo último ' - '
    parts = raw.rsplit(' - ', 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return raw, ''


def parse_sheet(ws):
    """Extrai cabeçalho e horários de uma aba com detecção automática de linhas."""
    # --- Cabeçalho ---
    header_row = detect_header_row(ws)
    header_raw = get_cell_value(ws, header_row, 1) or ''
    turma_text, prof_text = parse_turma_prof(header_raw)

    # --- Dados de aula ---
    data_start = detect_start_row(ws)
    
    horarios = []
    hora_idx = 0   # índice nos HORA_LABELS
    r = data_start

    while hora_idx < len(HORA_LABELS) and r <= data_start + 20:
        hora_label = HORA_LABELS[hora_idx]
        tipo_info = TIPOS_ESPECIAIS.get(hora_label)

        if tipo_info:
            # Linha especial (intervalo/almoço)
            tipo, nome = tipo_info
            horarios.append({"hora": hora_label, "tipo": tipo, "nome": nome})
            hora_idx += 1
            r += 1
            continue

        # Verifica se esta linha tem dados de aula (coluna A deve ter horário OU colunas B-F têm valores)
        col_a_val = get_cell_value(ws, r, 1)
        col_b_val = get_cell_value(ws, r, 2)

        # Pula linhas de SAÍDA / REUNIÃO — detecta pela col B (não pela col A que pode ter "14h40" no horário da aula)
        col_a_str = str(col_a_val or '').upper()
        col_b_str = str(col_b_val or '').upper()
        # Linha de saída: col A começa com "14H40" E col B tem "SAÍDA" ou "INTERVALO/SAÍDA"
        is_saida = ('14H40' in col_a_str and 'SA' in col_b_str)
        is_reuniao = 'REUNIÃO PEDAGÓGICA' in col_b_str or 'REUNIAO PEDAGOGICA' in col_b_str
        # Linhas de 15h, 16h em diante são sempre pós-aula
        is_pos_aula = ('15H' in col_a_str[:5] or '16H' in col_a_str[:5])
        if is_saida or is_reuniao or is_pos_aula:
            r += 1
            continue

        # Lê as aulas
        aulas = []
        for col in DAY_COLS:
            raw_val = get_cell_value(ws, r, col)
            aulas.append(clean_cell(raw_val))

        # Verifica se a linha está completamente vazia
        if all(a == '—' for a in aulas) and col_a_val is None:
            r += 1
            continue

        horarios.append({"hora": hora_label, "aulas": aulas})
        hora_idx += 1
        r += 1

    return turma_text, prof_text, horarios


# ---- Processamento principal ----
sheet_names = wb.sheetnames
data = {}

for tab_expected, js_key in ID_MAP.items():
    # Encontra a aba pelo nome (após strip)
    ws = None
    for sn in sheet_names:
        if sn.strip() == tab_expected.strip():
            ws = wb[sn]
            break

    if ws is None:
        print(f"NAO ENCONTRADA: '{tab_expected}'")
        continue

    turma_text, prof_text, horarios = parse_sheet(ws)
    data[js_key] = {
        "id_original": tab_expected.strip(),
        "turma": turma_text,
        "prof": prof_text,
        "dias": ["Segunda", "Terça", "Quarta", "Quinta", "Sexta"],
        "horarios": horarios
    }
    print(f"OK  {js_key:12s} | linhas={len(horarios)} | {turma_text[:45]}")


# ---- Template JS completo ----
JS_TEMPLATE = r"""
const turmasData = {DATA_JSON};

function getColorClass(aula, turmaId) {
    if (!aula || aula === '—') return '';
    const u = aula.toUpperCase();

    // Ed. Física — Marjorie (1º, 2º, 3º ano) vs Marcos (4º, 5º ano)
    if (u.includes('MARJORIE')) return 'ef-marjorie';
    if (u.includes('MARCOS'))   return 'ef-marcos';
    if (u.includes('ED. FÍS') || u.includes('ED. FIS') || u.includes('ED.FIS')) {
        if (turmaId && /^[123]/.test(turmaId)) return 'ef-marjorie';
        if (turmaId && /^[45]/.test(turmaId))  return 'ef-marcos';
        return 'ef-marjorie';
    }

    // Artes — Bianca (1º ano) vs Mirian (2º–5º ano)
    if (u.includes('BIANCA'))  return 'artes-bianca';
    if (u.includes('MIRIAN'))  return 'artes-mirian';
    if (u === 'A') {
        if (turma.startswith('1')) return 'artes-bianca';
        return 'artes-mirian';
    }

    if (u.includes('INGL')) return 'ingles';
    if (u.includes('LEITURA'))  return 'leitura';
    if (u.includes('MAKER'))    return 'maker';
    if (u.includes('SEBRAE'))   return 'sebrae';
    if (u.includes('PROERD'))   return 'proerd';
    if (u.includes('REUNIÃO') || u.includes('REUNIAO'))  return 'reuniao';
    if (u.includes('PEB'))      return 'peb1';
    return '';
}

function formatAulaCell(aula) {
    if (!aula || aula === '—') return '<span style="opacity:0.3">—</span>';

    // Separar disciplina de tipo (H.A. / ESTUDO / PARES / APOIO / REFORÇO)
    const tipoMatch = aula.match(/\s+(H\.A\.|ESTUDO|PARES|APOIO|REFORÇO)$/i);
    if (tipoMatch) {
        const disciplina = aula.slice(0, aula.length - tipoMatch[0].length).trim();
        const tipoStr    = tipoMatch[1].trim();
        const tipoClass  = tipoStr.toUpperCase() === 'H.A.' ? 'ha' : 'tipo';
        return '<span class="disciplina">' + disciplina + '</span><br><span class="tipo ' + tipoClass + '">' + tipoStr + '</span>';
    }
    return '<span class="disciplina">' + aula + '</span>';
}

function gerarTabelaHTML(turmaId) {
    const data = turmasData[turmaId];
    if (!data) return '<p style="color:red">Turma não encontrada: ' + turmaId + '</p>';

    let html = '<div class="tabela-horario">' +
        '<h2>' + data.turma + '</h2>' +
        '<h3>' + data.prof + '</h3>' +
        '<div class="table-responsive"><table><thead><tr>' +
        '<th>HORÁRIO</th>' +
        data.dias.map(d => '<th>' + d.toUpperCase() + '</th>').join('') +
        '</tr></thead><tbody>';

    data.horarios.forEach(linha => {
        if (linha.tipo) {
            html += '<tr class="linha-' + linha.tipo + '"><td>' + linha.hora + '</td><td colspan="5">' + linha.nome + '</td></tr>';
        } else {
            html += '<tr><td class="hora-col">' + linha.hora + '</td>' +
                linha.aulas.map(aula => '<td class="' + getColorClass(aula, turmaId) + '">' + formatAulaCell(aula) + '</td>').join('') +
                '</tr>';
        }
    });

    html += '</tbody></table></div></div>';
    return html;
}

function gerarTabelaGeralHTML() {
    const turmasIds = ['1A', '1B', '1C', '2A', '2B', '2C', '3A', '3B', '3C', '4A', '4B', '4C', '5A', '5B', '5C', '5D'];
    const diasSemana = ['SEGUNDA-FEIRA', 'TERÇA-FEIRA', 'QUARTA-FEIRA', 'QUINTA-FEIRA', 'SEXTA-FEIRA'];
    const aulasIndices = [0, 1, 3, 4, 5, 7, 8]; 

    function getAbreviacao(aula) {
        if (!aula || aula === '—') return '';
        const u = aula.toUpperCase();
        if (u.includes('PEB 1') || u.includes('PEB1') || u === 'PEB' || u.includes('INTERVALO') || u.includes('SAÍDA')) return '';
        if (u.includes('ED. FÍS') || u.includes('ED. FIS') || u.includes('ED.FIS')) {
            if (u.includes('PARES')) return 'PEF';
            return 'EF';
        }
        if (u.includes('INGL')) return 'I';
        if (u.includes('ARTES')) {
            if (u.includes('PARES')) return 'PAr';
            return 'A';
        }
        if (u.includes('MAKER')) return 'MK';
        if (u.includes('LEITURA')) return 'OL';
        if (u.includes('SEBRAE') || u.includes('DSE')) return 'DSE';
        if (u.includes('PROERD') || u.includes('LIMA')) return 'Lima';
        if (u.includes('PARES')) {
            return 'PAr';
        }
        return '';
    }

    let html = `
    <div class="tabela-horario consolidado-container">
        <div class="consolidado-header">
            <h1>CIEP JAGUARI - HORÁRIO 2026</h1>
            <div class="data-ref">DATA:25/02/2026</div>
        </div>
        <div class="table-responsive" style="margin-bottom: 20px;">
            <table class="tabela-consolidada">
                <thead>
                    <tr>
                        <th rowspan="2" class="col-fixa"></th>
                        `;
    
    for(let i=0; i<diasSemana.length; i++) {
        html += `<th colspan="7" class="dia-header">${diasSemana[i]}</th>`;
        if (i < diasSemana.length - 1) html += `<th class="spacer-col"></th>`;
    }
    html += `</tr><tr>`;
    
    for(let i=0; i<diasSemana.length; i++) {
        html += `<th>1ª</th><th>2ª</th><th>3ª</th><th>4ª</th><th>5ª</th><th>6ª</th><th>7ª</th>`;
        if (i < diasSemana.length - 1) html += `<th class="spacer-col"></th>`;
    }
    html += `</tr></thead><tbody>`;

    turmasIds.forEach(id => {
        const data = turmasData[id];
        if (!data) return;
        
        let rowHtml = `<tr><th class="col-fixa turma-label">${data.id_original}</th>`;
        
        for (let diaIdx = 0; diaIdx < 5; diaIdx++) {
            aulasIndices.forEach(horarioIdx => {
                const linha = data.horarios[horarioIdx];
                let aulaNome = '';
                if (linha && linha.aulas && linha.aulas[diaIdx]) {
                    aulaNome = linha.aulas[diaIdx];
                }
                const abrev = getAbreviacao(aulaNome);
                let colorClass = abrev ? getColorClass(aulaNome, id) : '';
                
                // Reuniões de pares são brancas com borda ou texto de cor
                if (abrev === 'PEF') {
                    rowHtml += `<td class="aula-cell reuni-pef">${abrev}</td>`;
                } else if (abrev === 'PAr') {
                    rowHtml += `<td class="aula-cell reuni-par">${abrev}</td>`;
                } else if (abrev === 'Lima') {
                    rowHtml += `<td class="aula-cell proerd-lima">${abrev}</td>`;
                } else {
                    rowHtml += `<td class="aula-cell ${colorClass}">${abrev}</td>`;
                }
            });
            if (diaIdx < 4) rowHtml += `<td class="spacer-col"></td>`;
        }
        rowHtml += `</tr>`;
        html += rowHtml;
    });

    html += `
                </tbody>
            </table>
        </div>
        
        <div class="legenda-section">
            <div class="legenda-box">
                <div class="legenda-title">LEGENDA</div>
                <div class="leg-item"><span class="color-box ef-marjorie">EF</span> Educação Física - Marjorie</div>
                <div class="leg-item"><span class="color-box ingles">I</span> Inglês - Marcelo</div>
                <div class="leg-item"><span class="color-box artes-bianca">A</span> Artes - Bianca</div>
                <div class="leg-item"><span class="color-box artes-mirian">A</span> Artes - Mirian</div>
                <div class="leg-item"><span class="color-box maker">MK</span> Of. Maker - Sirlene</div>
                <div class="leg-item"><span class="color-box sebrae">DS</span> Desenvolvimento SócioEmocional - Cherlane</div>
                <div class="leg-item"><span class="color-box leitura">OL</span> Oficina de Leitura e Expressão - Raquel Castelaneli</div>
                <div class="leg-item"><span class="color-box ef-marcos">EF</span> Educação Física - Marcos</div>
                <div class="leg-item"><span class="color-box proerd-lima">Lima</span> PROERD - Lima</div>
                <div class="leg-item"><span class="color-box reuni-pef">PEF</span> Reunião de Pares - Educação Física</div>
                <div class="leg-item"><span class="color-box reuni-par">PAr</span> Reunião de Pares - Artes</div>
            </div>
            
            <div class="legenda-box">
                <div class="legenda-title">REUNIÃO DE PARES</div>
                <p>1º ANO - Segunda - 3ª aula</p>
                <p>2º ANO - Quarta - 5ª aula</p>
                <p>3º ANO - Terça - 3ª aula</p>
                <p>4º ANO - Quinta - 3ª aula</p>
                <p>5º ANO - Segunda - 5ª aula</p>
                <p>ARTES - Quinta - 3ª aula</p>
                <p>Educação Física - Segunda - 6ª aula</p>
            </div>
            
            <div class="legenda-box salas-prof">
                <div class="legenda-title">SALAS E PROFESSORES</div>
                <div class="grid-2cols">
                    <div>
                        <p>1º A - Leonora</p>
                        <p>1º B - Cleusa</p>
                        <p>1º C - Gisleide</p>
                        <p>2º A - Meire Braz</p>
                        <p>2º B - Meire Canali</p>
                        <p>2º C - Cleide</p>
                        <p>3º A - Solange</p>
                        <p>3º B - Tayrine</p>
                        <p>3º C - Denise</p>
                    </div>
                    <div>
                        <p>4º A - Luciana</p>
                        <p>4º B - Patricia</p>
                        <p>4º C - Nilceia</p>
                        <p>5º A - Fabiana</p>
                        <p>5º B - Adriana</p>
                        <p>5º C - Fátima</p>
                        <p>5º D - Flávia</p>
                    </div>
                </div>
            </div>
        </div>
    </div>`;
    return html;
}

function mostrarTabela(id, tipo) {
    const container = document.getElementById('tabelas-container');

    if (tipo === 'turma') { const sp = document.getElementById('sel-prof'); if (sp) sp.value = ''; }
    if (tipo === 'prof')  { const st = document.getElementById('sel-turma'); if (st) st.value = ''; }

    if (!id) {
        container.innerHTML = '<div id="tabela-placeholder" class="empty-state"><i class="bi bi-search"></i><h3>Selecione uma opção</h3><p>Escolha uma turma ou professor para ver os horários</p></div>';
        return;
    }

    if (id === 'GERAL') {
        container.innerHTML = gerarTabelaGeralHTML();
        return;
    }

    container.innerHTML = gerarTabelaHTML(id);
}

// Auto-load via URL params
document.addEventListener('DOMContentLoaded', () => {
    const p = new URLSearchParams(window.location.search);
    const turmaParam = p.get('turma');
    const profParam  = p.get('prof');
    const tabParam   = p.get('tab');

    if (turmaParam) {
        const s = document.getElementById('sel-turma'); if (s) s.value = turmaParam;
        mostrarTabela(turmaParam, 'turma');
    } else if (profParam) {
        const s = document.getElementById('sel-prof'); if (s) s.value = profParam;
        mostrarTabela(profParam, 'prof');
    } else if (tabParam === 'geral') {
        mostrarTabela('GERAL', 'geral');
    }
});
"""

data_json = json.dumps(data, indent=4, ensure_ascii=False)
js_content = JS_TEMPLATE.replace('{DATA_JSON}', data_json)

out_path = r"c:\Users\Usuario1\Downloads\sistema-escolar-v2-main (2)\sistema-escolar-v2-main\direcao\horario-jaguari.js"
with open(out_path, "w", encoding="utf-8") as f:
    f.write(js_content)

print(f"\nGerado horario-jaguari.js com sucesso! ({len(data)} entradas)")
